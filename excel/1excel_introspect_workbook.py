#!/usr/bin/env python3
import sys
import openpyxl

'''
python 1excel_introspect_workbook.py ./sales_2013.xlsx
'''
input_file = sys.argv[1]

workbook = openpyxl.load_workbook(input_file)

print('Number of worksheets:', workbook.worksheets.count)
for worksheet in workbook.worksheets:
	print("Worksheet name:", worksheet.title, "\tRows:", \
			worksheet.max_row, "\tColumns:", worksheet.max_column)
