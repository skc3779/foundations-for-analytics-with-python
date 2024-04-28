#!/usr/bin/env python3
import sys
# from xlrd import open_workbook
# from xlwt import Workbook
import openpyxl
'''
python 2excel_parsing_and_write.py ./sales_2013.xlsx ./output_files/2output.xlsx
'''
input_file = sys.argv[1]
output_file = sys.argv[2]


workbook = openpyxl.load_workbook(input_file)
# 새로운 워크시트 생성
output_worksheet = workbook.create_sheet("jan_2013_output")

worksheet = workbook.get_sheet_by_name('january_2013')
for row in worksheet.iter_rows():
	for cell in row:
		output_worksheet.cell(cell.row, cell.column, cell.value)

workbook.save(output_file)

'''
output_workbook = Workbook()
output_worksheet = output_workbook.add_sheet('jan_2013_output')

with open_workbook(input_file) as workbook:
	worksheet = workbook.sheet_by_name('january_2013')
	for row_index in range(worksheet.nrows):
		for column_index in range(worksheet.ncols):
			output_worksheet.write(row_index, column_index, worksheet.cell_value(row_index, column_index))
output_workbook.save(output_file)
'''