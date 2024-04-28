#!/usr/bin/env python3
import sys
from datetime import date
# from xlrd import open_workbook, xldate_as_tuple
# from xlwt import Workbook
import openpyxl
from openpyxl.cell import Cell

'''
python 2excel_parsing_and_write.py ./sales_2013.xlsx ./output_files/3output.xlsx
'''

input_file = sys.argv[1]
output_file = sys.argv[2]

workbook = openpyxl.load_workbook(input_file)
# 새로운 워크시트 생성
output_worksheet = workbook.create_sheet("jan_2013_output")

worksheet = workbook.get_sheet_by_name('january_2013')
for row in range(2, worksheet.max_row + 1):
	row_list_output = []
	for col in range(2, worksheet.max_column + 1):
		print('cell type : ', type(worksheet.cell(row=row, column=col).value).__name__)
		if type(worksheet.cell(row=row, column=col).value).__name__ == 'datetime':
			date_cell = worksheet.cell(row=row, column=col).value
			date_cell = date_cell.strftime('%m/%d/%Y')
			row_list_output.append(date_cell)
			output_worksheet.cell(row=row, column=col, value=date_cell)
		else:
			non_date_cell: Cell = worksheet.cell(row=row, column=col).value
			row_list_output.append(non_date_cell)
			output_worksheet.cell(row=row, column=col, value=non_date_cell)

workbook.save(output_file)




'''
output_workbook = Workbook()
output_worksheet = output_workbook.add_sheet('jan_2013_output')

with open_workbook(input_file) as workbook:
	worksheet = workbook.sheet_by_name('january_2013')
	for row_index in range(worksheet.nrows):
		row_list_output = []
		for col_index in range(worksheet.ncols):
			if worksheet.cell_type(row_index, col_index) == 3:
				date_cell = xldate_as_tuple(worksheet.cell_value\
					(row_index, col_index),workbook.datemode)
				date_cell = date(*date_cell[0:3]).strftime\
					('%m/%d/%Y')
				row_list_output.append(date_cell)
				output_worksheet.write(row_index, col_index, date_cell)
			else:
				non_date_cell = worksheet.cell_value\
					(row_index,col_index)
				row_list_output.append(non_date_cell)
				output_worksheet.write(row_index, col_index,\
					non_date_cell)
output_workbook.save(output_file)
'''