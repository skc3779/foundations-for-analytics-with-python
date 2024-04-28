#!/usr/bin/env python3
import re
import sys
from datetime import date
# from xlrd import open_workbook, xldate_as_tuple
# from xlwt import Workbook
import openpyxl
import warnings

'''
python 2excel_parsing_and_write.py ./sales_2013.xlsx ./output_files/6output.xlsx
'''

warnings.filterwarnings('ignore')

input_file = sys.argv[1]
output_file = sys.argv[2]

pattern = re.compile(r'(?P<my_pattern>^J.*)')
customer_name_column_index = 1
workbook = openpyxl.load_workbook(input_file)

# 첫 번째 Sheet
worksheet = workbook['january_2013']
# 새로운 워크시트 생성
output_worksheet = workbook.create_sheet("jan_2013_output")
data = []
header = [cell.value for cell in worksheet[1]]
print(header)
data.append(header)
for row in worksheet.iter_rows(min_row=2):
    row_list = []
    customer_name = row[customer_name_column_index].value
    print('customer_name', customer_name)
    if pattern.search(customer_name) is not None:
        for col in row:
            cell_value = col.value
            cell_type = col.data_type
            if cell_type == 'd':
                date_cell = cell_value.strftime('%m/%d/%Y')
                row_list.append(date_cell)
            else:
                row_list.append(cell_value)
    if row_list:
        data.append(row_list)

for list_index, output_list in enumerate(data):
    for element_index, element in enumerate(output_list):
        # row, column 1 ~ 시작한다.
        output_worksheet.cell(row=list_index+1, column=element_index+1, value=element)

workbook.save(output_file)

'''

output_workbook = Workbook()
output_worksheet = output_workbook.add_sheet('jan_2013_output')

pattern = re.compile(r'(?P<my_pattern>^J.*)')

customer_name_column_index = 1
with open_workbook(input_file) as workbook:
    worksheet = workbook.sheet_by_name('january_2013')
    data = []
    header = worksheet.row_values(0)
    data.append(header)
    for row_index in range(1, worksheet.nrows):		
        row_list = []
        if pattern.search(worksheet.cell_value(row_index, customer_name_column_index)):
            for column_index in range(worksheet.ncols):
                cell_value = worksheet.cell_value(row_index,column_index)
                cell_type = worksheet.cell_type(row_index, column_index)
                if cell_type == 3:
                    date_cell = xldate_as_tuple(cell_value,workbook.datemode)
                    date_cell = date(*date_cell[0:3]).strftime('%m/%d/%Y')
                    row_list.append(date_cell)
                else:
                    row_list.append(cell_value)
        if row_list:
            data.append(row_list)

    for list_index, output_list in enumerate(data):
        for element_index, element in enumerate(output_list):
            output_worksheet.write(list_index, element_index, element)

output_workbook.save(output_file)
'''
